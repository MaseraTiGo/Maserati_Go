#-*- coding: utf-8 -*-  #not neccesary 4 py3,because py3 bult in
import os, time, re, sys #not pythonic, but I like
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color
from multiprocessing import Process, Queue, Lock, cpu_count
#import cProfile

class PicProcess(Process):
    '''
    multi processes pic2excel
    '''
    def __init__(self, pro_name, queue, lock):
        super(PicProcess, self).__init__()
        self.pro_name = pro_name
        self.queue = queue
        self.lock = lock
    
    def make_excel(self, path):
        img = Image.open(path)
        pixels = img.load()
        width, higth = img.size
        excel_name, kind = path.split('\\')[-1].split('.')
        wb = Workbook()
        ws = wb['Sheet']
        patterns = []
        for row in range(higth):
            for col in range(width):
                if re.search('png', kind, re.I):
                    r, g, b, s = pixels[col, row]
                else:
                    r, g, b = pixels[col, row]                
                pre_process = map(lambda x: hex(x), [r, g, b])
                hex_pixel = list(map(lambda x: x.split('x')[-1].zfill(2), pre_process))
                hex_pixel.insert(0, '00')
                hex_value = ''.join(hex_pixel)
                #print('color', r, g, b, hex_value)
                #sys.exit()
                base_color = Color(rgb=hex_value)
                color_pattern = PatternFill(patternType='solid', fgColor=base_color, bgColor=base_color)
                patterns.append(color_pattern)
        position = 0
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=higth, max_col=width):
            for cell in row:
                cell.fill = patterns[position]
                position += 1
        wb.save('d:\\%s.xlsx'%excel_name)
    
    def run(self):
        while 1:
            self.lock.acquire()
            try:
                path = self.queue.get(True, 1)
            except Exception as e:
                print('exception!!!', e)
                self.lock.release()
                sys.exit()
            self.lock.release()
            self.make_excel(path)
            print('current process %s done'%self.pro_name)

class ProcessPool:
    lock = Lock()
    queue = Queue()
    processes = []
    def __init__(self, max=4):
        for i in range(max):
            process = PicProcess(str(i), self.queue, self.lock)
            self.processes.append(process)
    
    def add_data(self, data):
        self.lock.acquire()
        self.queue.put(data)
        self.lock.release()
    
    #@staticmethod
    def start(self):
        for p in self.processes:
            p.start()
    
    #@staticmethod
    def join(self):
        for process in self.processes:
            process.join()
    
if __name__ == '__main__':
    #profile = cProfile.profile()
    #profile.enable()
    s = time.time()
    src_path = '.'
    if len(sys.argv) > 1:
        src_path = sys.argv[1]
        print('path is :', sys.argv[1])
    obj = os.listdir(src_path)
    pics = list(filter(lambda x: re.search('\.jpg|\.png', x, re.I), obj))
    pics_path = list(map(lambda x: os.sep.join([src_path, x]), pics))
    core = cpu_count()
    num = len(pics) if len(pics) < core else core
    print('num:', num)
    pool = ProcessPool(num)
    while 1:
        if pics_path:
            pool.add_data(pics_path.pop())
        else:
            break
    pool.start()
    pool.join()
    print('main process done, time cost:', time.time()-s)
    #profile.disable()
    #import pstats
    #pstats.Stats(profile).sort_stats('ctime').print_stats(15, 0.1, '*.*')
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    